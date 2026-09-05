"""Bounded OpenTopia XLSX mutation worker.

The Rust host owns validation, path authorization, staging, size limits, and
result serialization. This worker deliberately has a narrow contract: apply an
already-validated write request with openpyxl and save it to the staged output.
"""

from __future__ import annotations

import json
import io
import sys
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def fail(message: str) -> None:
    print(json.dumps({"ok": False, "error": message}), file=sys.stderr)
    raise SystemExit(1)


def find_sheet(workbook: Workbook, requested: str):
    requested_folded = requested.casefold()
    for worksheet in workbook.worksheets:
        if worksheet.title.casefold() == requested_folded:
            return worksheet
    return None


def open_workbook(path: str):
    suffix = Path(path).suffix.casefold()
    return load_workbook(
        path,
        data_only=False,
        keep_links=True,
        keep_vba=suffix in {".xlsm", ".xltm"},
    )


def decode_value(encoded: dict[str, Any]) -> Any:
    kind = encoded["type"]
    if kind == "blank":
        return None
    if kind in {"string", "integer", "number", "boolean"}:
        return encoded["value"]
    if kind == "formula":
        expression = encoded["value"]["expression"].strip()
        return expression if expression.startswith("=") else f"={expression}"
    fail(f"unsupported cell value type: {kind}")


def create_workbook(sheets: list[dict[str, Any]]) -> Workbook:
    if not sheets:
        fail("a workbook must contain at least one worksheet")
    workbook = Workbook()
    first = workbook.active
    first.title = sheets[0]["name"]
    for sheet in sheets[1:]:
        workbook.create_sheet(sheet["name"])
    return workbook


def copy_worksheet(source, destination) -> None:
    """Copy cell content and ordinary worksheet presentation across workbooks."""
    for row in source.iter_rows():
        for source_cell in row:
            destination_cell = destination.cell(source_cell.row, source_cell.column)
            destination_cell.value = source_cell.value
            if source_cell.has_style:
                destination_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                destination_cell.number_format = source_cell.number_format
            if source_cell.hyperlink:
                destination_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                destination_cell.comment = copy(source_cell.comment)
    for key, dimension in source.column_dimensions.items():
        destination.column_dimensions[key] = copy(dimension)
    for key, dimension in source.row_dimensions.items():
        destination.row_dimensions[key] = copy(dimension)
    for merged_range in source.merged_cells.ranges:
        destination.merge_cells(str(merged_range))
    destination.freeze_panes = source.freeze_panes
    destination.sheet_format = copy(source.sheet_format)
    destination.sheet_properties = copy(source.sheet_properties)
    destination.page_margins = copy(source.page_margins)
    destination.page_setup = copy(source.page_setup)
    destination.print_options = copy(source.print_options)
    if source.auto_filter.ref:
        destination.auto_filter.ref = source.auto_filter.ref


def delete_row_groups(worksheet, rows: list[int]) -> int:
    ordered = sorted(set(rows), reverse=True)
    deleted = 0
    index = 0
    while index < len(ordered):
        high = ordered[index]
        low = high
        index += 1
        while index < len(ordered) and ordered[index] == low - 1:
            low = ordered[index]
            index += 1
        worksheet.delete_rows(low + 1, high - low + 1)
        deleted += high - low + 1
    return deleted


def apply_structure_operations(workbook: Workbook, operations: list[dict[str, Any]]) -> int:
    applied = 0
    source_workbooks: dict[str, Workbook] = {}
    for operation in operations:
        kind = operation["type"]
        if kind == "copy_sheet":
            source_path = operation["source"]
            source_workbook = source_workbooks.get(source_path)
            if source_workbook is None:
                source_workbook = open_workbook(source_path)
                source_workbooks[source_path] = source_workbook
            source_sheet = find_sheet(source_workbook, operation["sourceSheet"])
            if source_sheet is None:
                fail(f"source sheet not found: {operation['sourceSheet']}")
            if find_sheet(workbook, operation["destinationSheet"]) is not None:
                fail(f"destination sheet already exists: {operation['destinationSheet']}")
            destination = workbook.create_sheet(operation["destinationSheet"])
            copy_worksheet(source_sheet, destination)
            visibility = operation.get("visibility")
            if visibility is not None:
                destination.sheet_state = {
                    "visible": "visible",
                    "hidden": "hidden",
                    "very_hidden": "veryHidden",
                }[visibility]
            applied += 1
        elif kind == "delete_rows":
            worksheet = find_sheet(workbook, operation["sheet"])
            if worksheet is None:
                fail(f"sheet not found: {operation['sheet']}")
            applied += delete_row_groups(worksheet, operation["rows"])
        elif kind == "delete_sheet":
            worksheet = find_sheet(workbook, operation["sheet"])
            if worksheet is None:
                fail(f"sheet not found: {operation['sheet']}")
            workbook.remove(worksheet)
            applied += 1
        elif kind == "set_number_format":
            worksheet = find_sheet(workbook, operation["sheet"])
            if worksheet is None:
                fail(f"sheet not found: {operation['sheet']}")
            cell_range = operation["range"]
            start = cell_range["start"]
            end = cell_range["end"]
            for row in worksheet.iter_rows(
                min_row=start["row"] + 1,
                max_row=end["row"] + 1,
                min_col=start["column"] + 1,
                max_col=end["column"] + 1,
            ):
                for cell in row:
                    cell.number_format = operation["numberFormat"]
            applied += 1
        else:
            fail(f"unsupported structure operation: {kind}")
    return applied


def main() -> None:
    # Rust serializes the request as UTF-8. Windows' default text code page may
    # be CP936, so sys.stdin must not decide how workbook and sheet names decode.
    request = json.load(io.TextIOWrapper(sys.stdin.buffer, encoding="utf-8"))
    source = request.get("source")
    output = Path(request["output"])
    sheets = request.get("sheets", [])
    workbook = (
        open_workbook(source)
        if source
        else create_workbook(sheets)
    )

    applied_structure_operations = apply_structure_operations(
        workbook, request.get("operations", [])
    )

    for sheet_request in sheets:
        worksheet = find_sheet(workbook, sheet_request["name"])
        if worksheet is None:
            worksheet = workbook.create_sheet(sheet_request["name"])
        visibility = sheet_request.get("visibility")
        if visibility is not None:
            worksheet.sheet_state = {
                "visible": "visible",
                "hidden": "hidden",
                "very_hidden": "veryHidden",
            }[visibility]
        for update in sheet_request.get("cells", []):
            address = update["address"]
            worksheet.cell(
                row=address["row"] + 1,
                column=address["column"] + 1,
            ).value = decode_value(update["value"])

    if not workbook.worksheets:
        fail("a workbook must contain at least one worksheet")
    if not any(sheet.sheet_state == "visible" for sheet in workbook.worksheets):
        fail("a workbook must contain at least one visible worksheet")

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(
        json.dumps(
            {
                "ok": True,
                "output": str(output),
                "appliedStructureOperations": applied_structure_operations,
            }
        )
    )


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception as error:  # The Rust host converts this to a typed write failure.
        fail(f"{type(error).__name__}: {error}")
